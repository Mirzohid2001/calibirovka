from django.shortcuts import render, get_object_or_404, redirect
from django.http import JsonResponse
from django.contrib import messages
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_http_methods
from django.core.paginator import Paginator
from .models import (
    Tank,
    Product,
    TransferCalculation,
    VolumeWeightCalculation,
    AddingCalculation,
    DensityTemperatureCalculation,
    GasolineBlendCalculation,
    SavedProductConfiguration,
)
import json
import logging
from decimal import Decimal
from itertools import combinations
from .optimization import optimize_multi_product_blend

logger = logging.getLogger(__name__)

TEMPERATURE_CORRECTION_TABLE = [
    # Диапазоны плотности при 20°C (г/см³) и средние температурные поправки на 1°C (г/см³)
    (0.6500, 0.6599, 0.000962),
    (0.6600, 0.6699, 0.000949),
    (0.6700, 0.6799, 0.000936),
    (0.6800, 0.6899, 0.000925),
    (0.6900, 0.6999, 0.000910),
    (0.7000, 0.7099, 0.000897),
    (0.7100, 0.7199, 0.000884),
    (0.7200, 0.7299, 0.000870),
    (0.7300, 0.7399, 0.000857),
    (0.7400, 0.7499, 0.000844),
    (0.7500, 0.7599, 0.000831),
    (0.7600, 0.7699, 0.000818),
    (0.7700, 0.7799, 0.000805),
    (0.7800, 0.7899, 0.000792),
    (0.7900, 0.7999, 0.000778),
    (0.8000, 0.8099, 0.000765),
    (0.8100, 0.8199, 0.000752),
    (0.8200, 0.8299, 0.000738),
    (0.8300, 0.8399, 0.000725),
    (0.8400, 0.8499, 0.000712),
    (0.8500, 0.8599, 0.000699),
    (0.8600, 0.8699, 0.000686),
    (0.8700, 0.8799, 0.000673),
    (0.8800, 0.8899, 0.000660),
    (0.8900, 0.8999, 0.000647),
    (0.9000, 0.9099, 0.000633),
    (0.9100, 0.9199, 0.000620),
    (0.9200, 0.9299, 0.000607),
    (0.9300, 0.9399, 0.000594),
    (0.9400, 0.9499, 0.000581),
    (0.9500, 0.9599, 0.000567),
    (0.9600, 0.9699, 0.000554),
    (0.9700, 0.9799, 0.000541),
    (0.9800, 0.9899, 0.000528),
    (0.9900, 1.0000, 0.000515),
]
DEFAULT_TEMPERATURE_CORRECTION = 0.00065  # г/см³ на °C


def get_temperature_correction(density_kg_m3):
    """
    Возвращает температурную поправку (кг/м³ на °C) согласно таблице dobmaster.ru/73.html.
    В таблице используются значения плотности и поправки в г/см³, поэтому выполняется конвертация.
    """
    try:
        density_g_cm3 = float(density_kg_m3) / 1000.0
    except (TypeError, ValueError):
        density_g_cm3 = None

    correction_g = DEFAULT_TEMPERATURE_CORRECTION

    if density_g_cm3 is not None:
        for lower, upper, correction in TEMPERATURE_CORRECTION_TABLE:
            if lower <= density_g_cm3 <= upper:
                correction_g = correction
                break

    # Переводим в кг/м³ (1 г/см³ = 1000 кг/м³)
    return correction_g * 1000.0


def normalize_density_input(value):
    """
    Приводит плотность к кг/м³, если пользователь ввел значение в кг/л (0.x-1.x).
    Возвращает кортеж (плотность_кг_м3, примечание).
    """
    if value <= 0:
        raise ValueError("Плотность должна быть положительным числом.")
    
    if value < 10:  # предполагаем, что значение введено в кг/л
        converted = value * 1000
        note = (
            f"Введенная плотность {value:g} была интерпретирована как кг/л и "
            f"преобразована в {converted:.1f} кг/м³."
        )
        return converted, note
    
    return value, None


def home(request):
    """Главная страница с калькулятором"""
    tanks = Tank.objects.all()
    products = Product.objects.all()
    
    if request.method == 'POST':
        try:
            # Получить данные из формы
            tank_id = request.POST.get('tank')
            product_id = request.POST.get('product')
            density_str = request.POST.get('density_kg_per_liter', '').replace(',', '.')
            initial_height_str = request.POST.get('initial_height_cm', '').replace(',', '.')
            transfer_weight_str = request.POST.get('transfer_weight_kg', '').replace(',', '.')
            
            # Валидация входных данных
            if not all([tank_id, product_id, density_str, initial_height_str, transfer_weight_str]):
                messages.error(request, "Пожалуйста, заполните все поля.")
                return render(request, 'calibration/home.html', {
                    'tanks': tanks,
                    'products': products
                })
            
            # Преобразование в числа
            try:
                density = float(density_str)
                initial_height = float(initial_height_str)
                transfer_weight = float(transfer_weight_str)
            except ValueError:
                messages.error(request, "Пожалуйста, введите корректные числовые значения.")
                return render(request, 'calibration/home.html', {
                    'tanks': tanks,
                    'products': products
                })
            
            # Валидация диапазонов
            if density <= 0 or density > 5:
                messages.error(request, "Плотность должна быть между 0.0001 и 5.0000 кг/л")
                return render(request, 'calibration/home.html', {
                    'tanks': tanks,
                    'products': products
                })
            
            if initial_height < 0:
                messages.error(request, "Начальная высота не может быть отрицательной")
                return render(request, 'calibration/home.html', {
                    'tanks': tanks,
                    'products': products
                })
            
            if transfer_weight <= 0:
                messages.error(request, "Вес перекачки должен быть положительным числом")
                return render(request, 'calibration/home.html', {
                    'tanks': tanks,
                    'products': products
                })
            
            # Получить объекты
            tank = get_object_or_404(Tank, id=tank_id)
            product = get_object_or_404(Product, id=product_id)
            
            # Проверить, что начальная высота не превышает высоту резервуара
            if initial_height > tank.height_cm:
                messages.error(request, f"Начальная высота не может превышать высоту резервуара ({tank.height_cm:.2f} см)")
                return render(request, 'calibration/home.html', {
                    'tanks': tanks,
                    'products': products
                })
            
            # Выполнить расчеты
            try:
                # 1. Преобразовать начальную высоту в объем
                initial_volume = tank.height_to_volume(initial_height)
                
                # 2. Рассчитать удаленный объем из веса и плотности
                volume_removed = transfer_weight / density
                
                # 3. Проверить, что у нас достаточно жидкости для удаления
                if volume_removed > initial_volume:
                    messages.error(request, f"Невозможно удалить {volume_removed:.2f} л: в резервуаре только {initial_volume:.2f} л")
                    return render(request, 'calibration/home.html', {
                        'tanks': tanks,
                        'products': products
                    })
                
                # 4. Рассчитать конечный объем (удаляем жидкость)
                final_volume = initial_volume - volume_removed
                
                # 5. Преобразовать конечный объем в высоту
                final_height = tank.volume_to_height(final_volume)
                
                # 6. Рассчитать процент заполнения
                fill_percentage = (final_volume / tank.capacity_liters) * 100
                
                # 7. Определить метод интерполяции
                calibration_points_count = tank.calibrations.count()
                interpolation_method = 'spline' if calibration_points_count >= 4 else 'linear'
                
                # 8. Сохранить расчет в базе данных (сохраняем удаленный объем как положительное значение)
                calculation = TransferCalculation.objects.create(
                    tank=tank,
                    product=product,
                    density_kg_per_liter=density,
                    initial_height_cm=initial_height,
                    transfer_weight_kg=transfer_weight,
                    initial_volume_liters=initial_volume,
                    final_volume_liters=final_volume,
                    volume_added_liters=volume_removed,  # сохраняем как положительное значение удаленного объема
                    final_height_cm=final_height,
                    fill_percentage=fill_percentage,
                    interpolation_method=interpolation_method
                )
                
                # 9. Подготовить результат для отображения
                result = {
                    'tank_name': tank.name,
                    'product_name': product.name,
                    'density': density,
                    'initial_height': initial_height,
                    'transfer_weight': transfer_weight,
                    'initial_volume': initial_volume,
                    'final_volume': final_volume,
                    'volume_removed': volume_removed,
                    'final_height': final_height,
                    'fill_percentage': fill_percentage,
                    'interpolation_method': interpolation_method
                }
                
                messages.success(request, "Расчет выполнен успешно!")
                
                return render(request, 'calibration/home.html', {
                    'tanks': tanks,
                    'products': products,
                    'result': result
                })
                
            except Exception as e:
                logger.error(f"Ошибка расчета: {str(e)}")
                messages.error(request, f"Ошибка при выполнении расчета: {str(e)}")
                return render(request, 'calibration/home.html', {
                    'tanks': tanks,
                    'products': products
                })
        
        except Exception as e:
            logger.error(f"Общая ошибка: {str(e)}")
            messages.error(request, "Произошла ошибка при обработке запроса.")
            return render(request, 'calibration/home.html', {
                'tanks': tanks,
                'products': products
            })
    
    return render(request, 'calibration/home.html', {
        'tanks': tanks,
        'products': products
    })


def history(request):
    """Страница истории расчетов"""
    # Получить все типы расчетов
    transfer_calculations = TransferCalculation.objects.all().order_by('-timestamp')
    volume_weight_calculations = VolumeWeightCalculation.objects.all().order_by('-timestamp')
    adding_calculations = AddingCalculation.objects.all().order_by('-timestamp')
    density_calculations = DensityTemperatureCalculation.objects.all().order_by('-timestamp')
    gasoline_blend_calculations = GasolineBlendCalculation.objects.all().order_by('-timestamp')
    
    # Объединить все расчеты и отсортировать по времени
    all_calculations = []
    
    for calc in transfer_calculations:
        all_calculations.append({
            'type': 'transfer',
            'object': calc,
            'timestamp': calc.timestamp,
            'tank_name': calc.tank_name,
            'product_name': calc.product_name,
            'description': f"Откачка: {calc.transfer_weight_kg:.2f} кг из {calc.initial_height_cm:.2f} см → {calc.final_height_cm:.2f} см"
        })
    
    for calc in volume_weight_calculations:
        all_calculations.append({
            'type': 'volume_weight',
            'object': calc,
            'timestamp': calc.timestamp,
            'tank_name': calc.tank_name,
            'product_name': calc.product_name,
            'description': f"Объем и вес: {calc.height_cm:.2f} см → {calc.volume_liters:.2f} л, {calc.weight_kg:.2f} кг"
        })
    
    for calc in adding_calculations:
        all_calculations.append({
            'type': 'adding',
            'object': calc,
            'timestamp': calc.timestamp,
            'tank_name': calc.tank_name,
            'product_name': calc.product_name,
            'description': f"Добавление: {calc.current_height_cm:.2f} см + {calc.amount_value:.2f} {'кг' if calc.amount_type == 'weight' else 'л'} → {calc.final_height_cm:.2f} см"
        })
    
    for calc in density_calculations:
        all_calculations.append({
            'type': 'density',
            'object': calc,
            'timestamp': calc.timestamp,
            'tank_name': '—',
            'product_name': calc.product_name,
            'description': f"Плотность: {calc.reference_density_kg_m3:.1f} кг/м³ при {calc.reference_temperature_c:.1f}°C → {calc.corrected_density_kg_m3:.1f} кг/м³"
        })
    
    for calc in gasoline_blend_calculations:
        variants_count = len(calc.blend_variants) if calc.blend_variants else 0
        all_calculations.append({
            'type': 'gasoline_blend',
            'object': calc,
            'timestamp': calc.timestamp,
            'tank_name': '—',
            'product_name': f"AI-{calc.target_octane}",
            'description': f"Смешивание бензина: {variants_count} вариантов, целевое октановое число: {calc.target_octane}"
        })
    
    # Сортировка по времени (новые сначала)
    all_calculations.sort(key=lambda x: x['timestamp'], reverse=True)
    
    # Пагинация
    paginator = Paginator(all_calculations, 10)  # 10 расчетов на страницу
    page_number = request.GET.get('page')
    calculations = paginator.get_page(page_number)
    
    return render(request, 'calibration/history.html', {
        'calculations': calculations,
        'is_paginated': calculations.has_other_pages(),
        'page_obj': calculations
    })


def delete_calculation(request, calculation_id):
    """Удалить расчет"""
    if request.method == 'POST':
        # Попробуем найти расчет во всех трех моделях
        calculation = None
        
        # Проверяем TransferCalculation
        try:
            calculation = TransferCalculation.objects.get(id=calculation_id)
        except TransferCalculation.DoesNotExist:
            pass
        
        # Проверяем VolumeWeightCalculation
        if not calculation:
            try:
                calculation = VolumeWeightCalculation.objects.get(id=calculation_id)
            except VolumeWeightCalculation.DoesNotExist:
                pass
        
        # Проверяем AddingCalculation
        if not calculation:
            try:
                calculation = AddingCalculation.objects.get(id=calculation_id)
            except AddingCalculation.DoesNotExist:
                pass
        
        # Проверяем GasolineBlendCalculation
        if not calculation:
            try:
                calculation = GasolineBlendCalculation.objects.get(id=calculation_id)
            except GasolineBlendCalculation.DoesNotExist:
                pass
        
        if calculation:
            calculation.delete()
            messages.success(request, "Расчет удален успешно.")
        else:
            messages.error(request, "Расчет не найден.")
    
    return redirect('calibration:history')


def interpolate_volume_from_height(reservoir, height_cm):
    """
    Interpolate volume from height using calibration data.
    Returns volume in liters.
    """
    height_cm = Decimal(str(height_cm))
    
    # Get calibration points for this reservoir
    calibration_points = CalibrationData.objects.filter(
        reservoir=reservoir
    ).order_by('height_cm')
    
    if not calibration_points.exists():
        raise ValueError(f"No calibration data found for reservoir {reservoir.name}")
    
    # Convert to list for easier manipulation
    points = list(calibration_points.values_list('height_cm', 'volume_liters'))
    
    # Check if height is exactly in our data
    for point_height, point_volume in points:
        if height_cm == point_height:
            return point_volume
    
    # Check if height is outside our calibration range
    if height_cm < points[0][0]:
        raise ValueError(f"Height {height_cm}cm is below minimum calibrated height {points[0][0]}cm")
    if height_cm > points[-1][0]:
        raise ValueError(f"Height {height_cm}cm is above maximum calibrated height {points[-1][0]}cm")
    
    # Find the two points to interpolate between
    for i in range(len(points) - 1):
        h1, v1 = points[i]
        h2, v2 = points[i + 1]
        
        if h1 <= height_cm <= h2:
            # Linear interpolation
            ratio = (height_cm - h1) / (h2 - h1)
            interpolated_volume = v1 + ratio * (v2 - v1)
            return interpolated_volume.quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
    
    raise ValueError("Unable to interpolate volume for given height")


def interpolate_height_from_volume(reservoir, volume_liters):
    """
    Interpolate height from volume using calibration data.
    Returns height in cm.
    """
    volume_liters = Decimal(str(volume_liters))
    
    # Get calibration points for this reservoir
    calibration_points = CalibrationData.objects.filter(
        reservoir=reservoir
    ).order_by('height_cm')
    
    if not calibration_points.exists():
        raise ValueError(f"No calibration data found for reservoir {reservoir.name}")
    
    # Convert to list for easier manipulation
    points = list(calibration_points.values_list('height_cm', 'volume_liters'))
    
    # Check if volume is exactly in our data
    for point_height, point_volume in points:
        if volume_liters == point_volume:
            return point_height
    
    # Check if volume is outside our calibration range
    if volume_liters < points[0][1]:
        raise ValueError(f"Volume {volume_liters}L is below minimum calibrated volume {points[0][1]}L")
    if volume_liters > points[-1][1]:
        raise ValueError(f"Volume {volume_liters}L is above maximum calibrated volume {points[-1][1]}L")
    
    # Find the two points to interpolate between
    for i in range(len(points) - 1):
        h1, v1 = points[i]
        h2, v2 = points[i + 1]
        
        if v1 <= volume_liters <= v2:
            # Linear interpolation
            ratio = (volume_liters - v1) / (v2 - v1)
            interpolated_height = h1 + ratio * (h2 - h1)
            return interpolated_height.quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
    
    raise ValueError("Unable to interpolate height for given volume")


@csrf_exempt
@require_http_methods(["POST"])
def calculate_transfer(request):
    """API endpoint для расчета перекачки (для AJAX запросов)"""
    try:
        data = json.loads(request.body)
        
        # Получить данные
        tank_id = data.get('tank_id')
        product_id = data.get('product_id')
        density = float(str(data.get('density_kg_per_liter')).replace(',', '.'))
        initial_height = float(str(data.get('initial_height_cm')).replace(',', '.'))
        transfer_weight = float(str(data.get('transfer_weight_kg')).replace(',', '.'))
        
        # Валидация
        if density <= 0 or density > 5:
            return JsonResponse({
                'success': False,
                'error': 'Плотность должна быть между 0.0001 и 5.0000 кг/л'
            })
        
        if initial_height < 0:
            return JsonResponse({
                'success': False,
                'error': 'Начальная высота не может быть отрицательной'
            })
        
        if transfer_weight <= 0:
            return JsonResponse({
                'success': False,
                'error': 'Вес перекачки должен быть положительным числом'
            })
        
        # Получить объекты
        tank = Tank.objects.get(id=tank_id)
        product = Product.objects.get(id=product_id)
        
        # Проверить высоту резервуара
        if initial_height > tank.height_cm:
            return JsonResponse({
                'success': False,
                'error': f'Начальная высота не может превышать высоту резервуара ({tank.height_cm:.2f} см)'
            })
        
        # Выполнить расчеты
        initial_volume = tank.height_to_volume(initial_height)
        volume_removed = transfer_weight / density
        
        # Проверить, что у нас достаточно жидкости для удаления
        if volume_removed > initial_volume:
            return JsonResponse({
                'success': False,
                'error': f'Невозможно удалить {volume_removed:.2f} л: в резервуаре только {initial_volume:.2f} л'
            })
        
        final_volume = initial_volume - volume_removed
        final_height = tank.volume_to_height(final_volume)
        fill_percentage = (final_volume / tank.capacity_liters) * 100
        
        # Определить метод интерполяции
        interpolation_method = 'spline' if tank.calibrations.count() >= 4 else 'linear'
        
        # Сохранить расчет
        TransferCalculation.objects.create(
            tank=tank,
            product=product,
            density_kg_per_liter=density,
            initial_height_cm=initial_height,
            transfer_weight_kg=transfer_weight,
            initial_volume_liters=initial_volume,
            final_volume_liters=final_volume,
            volume_added_liters=volume_removed,  # сохраняем как положительное значение удаленного объема
            final_height_cm=final_height,
            fill_percentage=fill_percentage,
            interpolation_method=interpolation_method
        )
        
        return JsonResponse({
            'success': True,
            'tank_name': tank.name,
            'product_name': product.name,
            'density': density,
            'initial_height': initial_height,
            'transfer_weight': transfer_weight,
            'initial_volume': initial_volume,
            'final_volume': final_volume,
            'volume_removed': volume_removed,
            'final_height': final_height,
            'fill_percentage': fill_percentage,
            'interpolation_method': interpolation_method
        })
        
    except Exception as e:
        logger.error(f"API ошибка: {str(e)}")
        return JsonResponse({
            'success': False,
            'error': f'Ошибка при выполнении расчета: {str(e)}'
        })


def volume_weight_calculator(request):
    """Калькулятор объема и веса"""
    tanks = Tank.objects.all()
    products = Product.objects.all()
    
    if request.method == 'POST':
        try:
            # Получить данные из формы
            tank_id = request.POST.get('tank')
            product_id = request.POST.get('product')
            density_str = request.POST.get('density_kg_per_liter', '').replace(',', '.')
            height_str = request.POST.get('height_cm', '').replace(',', '.')
            
            # Валидация входных данных
            if not all([tank_id, product_id, density_str, height_str]):
                messages.error(request, "Пожалуйста, заполните все поля.")
                return render(request, 'calibration/volume_weight.html', {
                    'tanks': tanks,
                    'products': products
                })
            
            # Преобразование в числа
            try:
                density = float(density_str)
                height = float(height_str)
            except ValueError:
                messages.error(request, "Пожалуйста, введите корректные числовые значения.")
                return render(request, 'calibration/volume_weight.html', {
                    'tanks': tanks,
                    'products': products
                })
            
            # Валидация диапазонов
            if density <= 0 or density > 5:
                messages.error(request, "Плотность должна быть между 0.0001 и 5.0000 кг/л")
                return render(request, 'calibration/volume_weight.html', {
                    'tanks': tanks,
                    'products': products
                })
            
            if height < 0:
                messages.error(request, "Высота не может быть отрицательной")
                return render(request, 'calibration/volume_weight.html', {
                    'tanks': tanks,
                    'products': products
                })
            
            # Получить объекты
            tank = get_object_or_404(Tank, id=tank_id)
            product = get_object_or_404(Product, id=product_id)
            
            # Проверить, что высота не превышает высоту резервуара
            if height > tank.height_cm:
                messages.error(request, f"Высота не может превышать высоту резервуара ({tank.height_cm:.2f} см)")
                return render(request, 'calibration/volume_weight.html', {
                    'tanks': tanks,
                    'products': products
                })
            
            # Выполнить расчеты
            try:
                # 1. Преобразовать высоту в объем
                volume = tank.height_to_volume(height)
                
                # 2. Рассчитать вес (объем × плотность)
                weight = volume * density
                
                # 3. Рассчитать процент заполнения
                fill_percentage = (volume / tank.capacity_liters) * 100
                
                # 4. Определить метод интерполяции
                calibration_points_count = tank.calibrations.count()
                interpolation_method = 'spline' if calibration_points_count >= 4 else 'linear'
                
                # 5. Сохранить расчет в базе данных
                calculation = VolumeWeightCalculation.objects.create(
                    tank=tank,
                    product=product,
                    height_cm=height,
                    density_kg_per_liter=density,
                    volume_liters=volume,
                    weight_kg=weight,
                    fill_percentage=fill_percentage,
                    interpolation_method=interpolation_method
                )
                
                # 6. Подготовить результат для отображения
                result = {
                    'tank_name': tank.name,
                    'product_name': product.name,
                    'height': height,
                    'density': density,
                    'volume': volume,
                    'weight': weight,
                    'fill_percentage': fill_percentage,
                    'interpolation_method': interpolation_method
                }
                
                messages.success(request, "Расчет выполнен успешно!")
                
                return render(request, 'calibration/volume_weight.html', {
                    'tanks': tanks,
                    'products': products,
                    'result': result
                })
                
            except Exception as e:
                logger.error(f"Ошибка расчета: {str(e)}")
                messages.error(request, f"Ошибка при выполнении расчета: {str(e)}")
                return render(request, 'calibration/volume_weight.html', {
                    'tanks': tanks,
                    'products': products
                })
        
        except Exception as e:
            logger.error(f"Общая ошибка: {str(e)}")
            messages.error(request, "Произошла ошибка при обработке запроса.")
            return render(request, 'calibration/volume_weight.html', {
                'tanks': tanks,
                'products': products
            })
    
    return render(request, 'calibration/volume_weight.html', {
        'tanks': tanks,
        'products': products
    })


def adding_calculator(request):
    """Калькулятор добавления жидкости"""
    tanks = Tank.objects.all()
    products = Product.objects.all()
    
    if request.method == 'POST':
        try:
            # Получить данные из формы
            tank_id = request.POST.get('tank')
            product_id = request.POST.get('product')
            density_str = request.POST.get('density_kg_per_liter', '').replace(',', '.')
            current_height_str = request.POST.get('current_height_cm', '').replace(',', '.')
            amount_type = request.POST.get('amount_type')
            amount_value_str = request.POST.get('amount_value', '').replace(',', '.')
            
            # Валидация входных данных
            if not all([tank_id, product_id, density_str, current_height_str, amount_type, amount_value_str]):
                messages.error(request, "Пожалуйста, заполните все поля.")
                return render(request, 'calibration/adding.html', {
                    'tanks': tanks,
                    'products': products
                })
            
            # Преобразование в числа
            try:
                density = float(density_str)
                current_height = float(current_height_str)
                amount_value = float(amount_value_str)
            except ValueError:
                messages.error(request, "Пожалуйста, введите корректные числовые значения.")
                return render(request, 'calibration/adding.html', {
                    'tanks': tanks,
                    'products': products
                })
            
            # Валидация диапазонов
            if density <= 0 or density > 5:
                messages.error(request, "Плотность должна быть между 0.0001 и 5.0000 кг/л")
                return render(request, 'calibration/adding.html', {
                    'tanks': tanks,
                    'products': products
                })
            
            if current_height < 0:
                messages.error(request, "Текущая высота не может быть отрицательной")
                return render(request, 'calibration/adding.html', {
                    'tanks': tanks,
                    'products': products
                })
            
            if amount_value <= 0:
                messages.error(request, "Количество для добавления должно быть положительным")
                return render(request, 'calibration/adding.html', {
                    'tanks': tanks,
                    'products': products
                })
            
            # Получить объекты
            tank = get_object_or_404(Tank, id=tank_id)
            product = get_object_or_404(Product, id=product_id)
            
            # Проверить, что текущая высота не превышает высоту резервуара
            if current_height > tank.height_cm:
                messages.error(request, f"Текущая высота не может превышать высоту резервуара ({tank.height_cm:.2f} см)")
                return render(request, 'calibration/adding.html', {
                    'tanks': tanks,
                    'products': products
                })
            
            # Выполнить расчеты
            try:
                # 1. Преобразовать текущую высоту в объем
                current_volume = tank.height_to_volume(current_height)
                
                # 2. Рассчитать текущий вес
                current_weight = current_volume * density
                
                # 3. Рассчитать добавляемые объем и вес
                if amount_type == 'weight':
                    added_weight = amount_value
                    added_volume = added_weight / density
                else:  # volume
                    added_volume = amount_value
                    added_weight = added_volume * density
                
                # 4. Рассчитать конечные значения
                final_volume = current_volume + added_volume
                final_weight = current_weight + added_weight
                
                # 5. Проверить, не превышает ли конечный объем емкость резервуара
                if final_volume > tank.capacity_liters:
                    messages.error(request, f"Невозможно добавить {added_volume:.2f} л: превысит емкость резервуара ({tank.capacity_liters:.2f} л)")
                    return render(request, 'calibration/adding.html', {
                        'tanks': tanks,
                        'products': products
                    })
                
                # 6. Преобразовать конечный объем в высоту
                final_height = tank.volume_to_height(final_volume)
                
                # 7. Рассчитать процент заполнения
                fill_percentage = (final_volume / tank.capacity_liters) * 100
                
                # 8. Определить метод интерполяции
                calibration_points_count = tank.calibrations.count()
                interpolation_method = 'spline' if calibration_points_count >= 4 else 'linear'
                
                # 9. Сохранить расчет в базе данных
                calculation = AddingCalculation.objects.create(
                    tank=tank,
                    product=product,
                    current_height_cm=current_height,
                    density_kg_per_liter=density,
                    amount_type=amount_type,
                    amount_value=amount_value,
                    current_volume_liters=current_volume,
                    current_weight_kg=current_weight,
                    added_volume_liters=added_volume,
                    added_weight_kg=added_weight,
                    final_volume_liters=final_volume,
                    final_weight_kg=final_weight,
                    final_height_cm=final_height,
                    fill_percentage=fill_percentage,
                    interpolation_method=interpolation_method
                )
                
                # 10. Подготовить результат для отображения
                result = {
                    'tank_name': tank.name,
                    'product_name': product.name,
                    'current_height': current_height,
                    'density': density,
                    'amount_type': amount_type,
                    'amount_value': amount_value,
                    'current_volume': current_volume,
                    'current_weight': current_weight,
                    'added_volume': added_volume,
                    'added_weight': added_weight,
                    'final_volume': final_volume,
                    'final_weight': final_weight,
                    'final_height': final_height,
                    'fill_percentage': fill_percentage,
                    'interpolation_method': interpolation_method
                }
                
                messages.success(request, "Расчет выполнен успешно!")
                
                return render(request, 'calibration/adding.html', {
                    'tanks': tanks,
                    'products': products,
                    'result': result
                })
                
            except Exception as e:
                logger.error(f"Ошибка расчета: {str(e)}")
                messages.error(request, f"Ошибка при выполнении расчета: {str(e)}")
                return render(request, 'calibration/adding.html', {
                    'tanks': tanks,
                    'products': products
                })
        
        except Exception as e:
            logger.error(f"Общая ошибка: {str(e)}")
            messages.error(request, "Произошла ошибка при обработке запроса.")
            return render(request, 'calibration/adding.html', {
                'tanks': tanks,
                'products': products
            })
    
    return render(request, 'calibration/adding.html', {
        'tanks': tanks,
        'products': products
    })


def density_calculator(request):
    """Калькулятор корректировки плотности по температуре"""
    if request.method == 'POST':
        try:
            reference_density_str = request.POST.get('reference_density', '').replace(',', '.')
            reference_temperature_str = request.POST.get('reference_temperature', '').replace(',', '.')
            target_temperature_str = request.POST.get('target_temperature', '').replace(',', '.')
            notes = request.POST.get('notes', '').strip()

            if not all([reference_density_str, reference_temperature_str, target_temperature_str]):
                messages.error(request, "Пожалуйста, заполните все обязательные поля.")
                return render(request, 'calibration/density.html', {
                })

            try:
                reference_density_input = float(reference_density_str)
                reference_temperature = float(reference_temperature_str)
                target_temperature = float(target_temperature_str)
            except ValueError:
                messages.error(request, "Пожалуйста, введите корректные числовые значения.")
                return render(request, 'calibration/density.html', {
                })
            
            if reference_density_input <= 0:
                messages.error(request, "Плотность должна быть положительным числом.")
                return render(request, 'calibration/density.html', {
                })

            reference_density, density_note = normalize_density_input(reference_density_input)
            temperature_correction = get_temperature_correction(reference_density)

            delta_t = target_temperature - reference_temperature
            corrected_density = reference_density - temperature_correction * delta_t
            density_diff = corrected_density - reference_density

            calculation = DensityTemperatureCalculation.objects.create(
                product=None,
                reference_density_kg_m3=reference_density,
                reference_temperature_c=reference_temperature,
                target_temperature_c=target_temperature,
                thermal_expansion_coefficient=temperature_correction,
                corrected_density_kg_m3=corrected_density,
                density_difference_kg_m3=density_diff,
                notes=notes or None
            )

            result = {
                'reference_density': reference_density,
                'reference_temperature': reference_temperature,
                'target_temperature': target_temperature,
                'temperature_correction': temperature_correction,
                'corrected_density': corrected_density,
                'density_difference': density_diff,
                'notes': notes,
                'density_note': density_note,
            }

            messages.success(request, "Плотность успешно пересчитана!")

            return render(request, 'calibration/density.html', {
                'result': result
            })

        except Exception as e:
            logger.error(f"Ошибка расчета плотности: {str(e)}")
            messages.error(request, "Произошла ошибка при выполнении расчета.")

    return render(request, 'calibration/density.html', {
    })


def density_quick_calculator(request):
    """
    Упрощенный калькулятор пересчета плотности, требующий только фактическую плотность,
    текущую и целевую температуры. Использует типичный коэффициент 0.00065 1/°C.
    """
    result = None
    
    if request.method == 'POST':
        actual_density_str = request.POST.get('actual_density', '').replace(',', '.')
        actual_temp_str = request.POST.get('actual_temperature', '').replace(',', '.')
        desired_temp_str = request.POST.get('desired_temperature', '').replace(',', '.')
        
        if not all([actual_density_str, actual_temp_str, desired_temp_str]):
            messages.error(request, "Пожалуйста, заполните все поля формы.")
        else:
            try:
                actual_density_input = float(actual_density_str)
                actual_temp = float(actual_temp_str)
                desired_temp = float(desired_temp_str)
            except ValueError:
                messages.error(request, "Пожалуйста, введите корректные числовые значения.")
            else:
                if actual_density_input <= 0:
                    messages.error(request, "Плотность должна быть положительным числом.")
                else:
                    actual_density, density_note = normalize_density_input(actual_density_input)
                    temperature_correction = get_temperature_correction(actual_density)

                    delta_t = desired_temp - actual_temp
                    corrected_density = actual_density - temperature_correction * delta_t
                    density_diff = corrected_density - actual_density
                    
                    DensityTemperatureCalculation.objects.create(
                        product=None,
                        reference_density_kg_m3=actual_density,
                        reference_temperature_c=actual_temp,
                        target_temperature_c=desired_temp,
                        thermal_expansion_coefficient=temperature_correction,
                        corrected_density_kg_m3=corrected_density,
                        density_difference_kg_m3=density_diff,
                        notes="Быстрый калькулятор плотности"
                    )
                    
                    result = {
                        'actual_density': actual_density,
                        'actual_temperature': actual_temp,
                        'desired_temperature': desired_temp,
                        'corrected_density': corrected_density,
                        'density_difference': density_diff,
                        'temperature_correction': temperature_correction,
                        'density_note': density_note,
                    }
                    messages.success(request, "Плотность успешно пересчитана.")
    
    return render(request, 'calibration/density_quick.html', {
        'result': result
    })


# ============================================================
# БЕНЗИН АРАЛАШМА КАЛЬКУЛЯТОР ФУНКЦИЯЛАРИ
# ============================================================

def calculate_octane_blend(products_percentages):
    """
    Aralashmaning yakuniy oktan sonini hisoblaydi (linear mixing)
    
    Args:
        products_percentages: List of tuples [(product, percentage), ...]
    
    Returns:
        float: Yakuniy oktan soni
    """
    if not products_percentages:
        return 0.0
    
    total_octane = 0.0
    for product, percentage in products_percentages:
        if product.octane_number:
            total_octane += product.octane_number * (percentage / 100.0)
    
    return round(total_octane, 2)


def calculate_price_blend(products_percentages):
    """
    Aralashmaning yakuniy narxini hisoblaydi
    
    Args:
        products_percentages: List of tuples [(product, percentage), ...]
    
    Returns:
        float: Yakuniy narx so'm/litr
    """
    if not products_percentages:
        return 0.0
    
    total_price = Decimal('0.0')
    for product, percentage in products_percentages:
        if product.price_per_liter:
            total_price += Decimal(str(product.price_per_liter)) * Decimal(str(percentage)) / Decimal('100.0')
    
    return float(total_price)


def check_gost_compliance(products_percentages):
    """
    GOST talablariga mos kelishini tekshiradi
    
    Args:
        products_percentages: List of tuples [(product, percentage), ...]
    
    Returns:
        tuple: (is_compliant: bool, warnings: list)
    """
    warnings = []
    
    for product, percentage in products_percentages:
        if product.gost_percentage is not None:
            if percentage > product.gost_percentage:
                warnings.append(
                    f"{product.name}: {percentage:.1f}% > GOST {product.gost_percentage:.1f}%"
                )
    
    is_compliant = len(warnings) == 0
    return is_compliant, warnings


def calculate_two_product_blend(product1, product2, target_octane):
    """
    Ikki product uchun aralashma hisoblaydi
    
    Args:
        product1: Past oktanli product
        product2: Yuqori oktanli product
        target_octane: Maqsad oktan soni
    
    Returns:
        dict: {product1_percentage, product2_percentage} yoki None
    """
    oct1 = product1.octane_number if product1.octane_number else 0
    oct2 = product2.octane_number if product2.octane_number else 0
    
    if oct1 >= oct2:
        return None
    
    if not (oct1 <= target_octane <= oct2):
        return None
    
    # Formula: target = (oct1 * p1) + (oct2 * p2)
    # p1 + p2 = 100%
    # p2 = (target - oct1) / (oct2 - oct1) * 100
    try:
        p2 = ((target_octane - oct1) / (oct2 - oct1)) * 100
        p1 = 100 - p2
        
        if 0 <= p1 <= 100 and 0 <= p2 <= 100:
            return {
                'product1_percentage': round(p1, 2),
                'product2_percentage': round(p2, 2)
            }
    except ZeroDivisionError:
        pass
    
    return None


def find_blend_variants(target_octane, products_data, max_variants=20, total_volume=None):
    """
    Maqsad oktan soni uchun aralashma variantlarini topadi
    
    Args:
        target_octane: Maqsad oktan soni
        products_data: Dict {product_id: {'octane': int, 'price': float, 'gost_percentage': float}}
        max_variants: Maksimal variantlar soni
        total_volume: Ixtiyoriy: umumiy og'irlik kg da
    
    Returns:
        list: Variantlar ro'yxati (narx bo'yicha tartiblangan)
    """
    # 1. Productlarni filterlash va tayyorlash
    products_list = []
    for product_id, data in products_data.items():
        try:
            product = Product.objects.get(id=int(product_id), is_for_blending=True)
            octane = float(data.get('octane', 0))
            price = float(data.get('price', 0))
            gost_percentage = float(data.get('gost_percentage', 100))
            
            if octane > 0 and price > 0:
                # Product ma'lumotlarini o'zgartirish
                product.octane_number = octane
                product.price_per_liter = Decimal(str(price))
                product.gost_percentage = gost_percentage
                products_list.append(product)
        except (Product.DoesNotExist, ValueError, TypeError):
            continue
    
    if len(products_list) < 2:
        return []
    
    variants = []
    
    # 2. Past va yuqori oktanli productlarni ajratish
    products_lower = [p for p in products_list if p.octane_number < target_octane]
    products_upper = [p for p in products_list if p.octane_number > target_octane]
    products_equal = [p for p in products_list if p.octane_number == target_octane]
    
    # 3. To'g'ridan-to'g'ri oktan soniga mos productlar
    if products_equal:
        for product in products_equal:
            variant = {
                'variant_number': 0,  # Keyin raqamlanadi
                'products': [{
                    'product_id': product.id,
                    'product_name': product.name,
                    'octane': float(product.octane_number),
                    'percentage': 100.0,
                    'weight_kg': float(total_volume) if total_volume else None,
                    'price_per_kg': float(product.price_per_liter)
                }],
                'final_octane': float(product.octane_number),
                'final_price_per_kg': float(product.price_per_liter),
                'total_price': float(float(product.price_per_liter) * float(total_volume)) if total_volume else None,
                'gost_compliant': True,
                'gost_warnings': []
            }
            variants.append(variant)
    
    # 4. Ikki product kombinatsiyalari (past va yuqori oktanli)
    if products_lower and products_upper:
        for lower in products_lower:
            for upper in products_upper:
                blend = calculate_two_product_blend(lower, upper, target_octane)
                if blend:
                    p1_perc = blend['product1_percentage']
                    p2_perc = blend['product2_percentage']
                    
                    # GOST tekshiruvi
                    products_perc = [
                        (lower, p1_perc),
                        (upper, p2_perc)
                    ]
                    gost_compliant, gost_warnings = check_gost_compliance(products_perc)
                    
                    final_octane = calculate_octane_blend(products_perc)
                    final_price = calculate_price_blend(products_perc)
                    
                    variant = {
                        'variant_number': 0,
                        'products': [
                            {
                                'product_id': lower.id,
                                'product_name': lower.name,
                                'octane': float(lower.octane_number),
                                'percentage': p1_perc,
                                'weight_kg': float(float(total_volume) * float(p1_perc) / 100) if total_volume else None,
                                'price_per_kg': float(lower.price_per_liter)
                            },
                            {
                                'product_id': upper.id,
                                'product_name': upper.name,
                                'octane': float(upper.octane_number),
                                'percentage': p2_perc,
                                'weight_kg': float(float(total_volume) * float(p2_perc) / 100) if total_volume else None,
                                'price_per_kg': float(upper.price_per_liter)
                            }
                        ],
                        'final_octane': round(final_octane, 2),
                        'final_price_per_kg': round(final_price, 2),
                        'total_price': round(final_price * float(total_volume), 2) if total_volume else None,
                        'gost_compliant': gost_compliant,
                        'gost_warnings': gost_warnings
                    }
                    variants.append(variant)
    
    # 5. 3+ product kombinatsiyalari - optimizatsiya (ASOSIY)
    # Barcha 3+ product kombinatsiyalarini optimizatsiya qilamiz
    if len(products_list) >= 3:
        try:
            logger.info(f"Optimizatsiya boshlanmoqda: {len(products_list)} ta product, maqsad oktan: {target_octane}")
            optimized_variants = optimize_multi_product_blend(
                products_list, 
                target_octane, 
                max_products=min(4, len(products_list)), 
                step=0.5,
                use_ai=True,
                num_variants=10  # Ko'proq variantlar topish uchun
            )
            logger.info(f"Optimizatsiya natijasi: {len(optimized_variants)} ta variant topildi")
            if optimized_variants:
                logger.info(f"Variantlarning narxlari: {[v.get('final_price', 0) for v in optimized_variants[:5]]}")
            
            for opt_var in optimized_variants:
                if not opt_var or not opt_var.get('final_octane'):
                    continue
                
                products_perc = [
                    (opt_var['products'][i], opt_var['percentages'][i])
                    for i in range(len(opt_var['products']))
                ]
                
                gost_compliant, gost_warnings = check_gost_compliance(products_perc)
                final_octane = opt_var['final_octane']
                final_price = opt_var['final_price']
                octane_diff = abs(final_octane - target_octane)
                
                # Oktan farqi cheklovi - eng ko'pi bilan 50.0 gacha qabul qilamiz
                # (maqsad diapazondan tashqarida bo'lsa ham, eng yaqin variantlarni ko'rsatamiz)
                if octane_diff > 50.0:
                    continue
                
                # Variant yaratish
                variant_products = []
                # Genetic Algorithm natijasida faqat active productlar bo'lishi mumkin
                products_list = opt_var['products']
                percentages_list = opt_var['percentages']
                
                for i, product in enumerate(products_list):
                    if i >= len(percentages_list):
                        continue
                    pct = percentages_list[i]
                    if pct <= 0.01:  # Juda kichik foizlarni o'tkazib yuboramiz
                        continue
                    variant_products.append({
                        'product_id': product.id,
                        'product_name': product.name,
                        'octane': float(product.octane_number),
                        'percentage': round(pct, 2),
                        'weight_kg': float(float(total_volume) * float(pct) / 100) if total_volume else None,
                        'price_per_kg': float(product.price_per_liter)
                    })
                
                # Kamida 2 ta product bo'lishi kerak
                if len(variant_products) < 2:
                    continue
                
                variant = {
                    'variant_number': 0,
                    'products': variant_products,
                    'final_octane': round(final_octane, 2),
                    'final_price_per_kg': round(final_price, 2),
                    'total_price': round(final_price * float(total_volume), 2) if total_volume else None,
                    'gost_compliant': gost_compliant,
                    'gost_warnings': gost_warnings
                }
                variants.append(variant)
        except Exception as e:
            logger.error(f"Optimizatsiya xatoligi: {str(e)}")
            # Xatolik bo'lsa ham, davom etamiz
    
    # 5b. Barcha product kombinatsiyalari (2, 3, 4+ product) - fallback
    # Agar optimizatsiya variantlar topolmasa, oddiy kombinatsiyalarni sinab ko'ramiz
    if not variants and len(products_list) >= 2:
        # 2 ta product kombinatsiyalari (barcha juftlar)
        for i in range(len(products_list)):
            for j in range(i + 1, len(products_list)):
                p1 = products_list[i]
                p2 = products_list[j]
                
                # Oktanlarni tekshirish
                oct1 = float(p1.octane_number)
                oct2 = float(p2.octane_number)
                
                # Agar past va yuqori oktanli productlar bo'lsa
                if oct1 < target_octane < oct2 or oct2 < target_octane < oct1:
                    # To'g'ri tartibda
                    if oct1 > oct2:
                        p1, p2 = p2, p1
                        oct1, oct2 = oct2, oct1
                    
                    blend = calculate_two_product_blend(p1, p2, target_octane)
                    if blend:
                        p1_perc = blend['product1_percentage']
                        p2_perc = blend['product2_percentage']
                        
                        products_perc = [(p1, p1_perc), (p2, p2_perc)]
                        gost_compliant, gost_warnings = check_gost_compliance(products_perc)
                        final_octane = calculate_octane_blend(products_perc)
                        final_price = calculate_price_blend(products_perc)
                        
                        variant = {
                            'variant_number': 0,
                            'products': [
                                {
                                    'product_id': p1.id,
                                    'product_name': p1.name,
                                    'octane': float(p1.octane_number),
                                    'percentage': p1_perc,
                                    'weight_kg': float(float(total_volume) * float(p1_perc) / 100) if total_volume else None,
                                    'price_per_kg': float(p1.price_per_liter)
                                },
                                {
                                    'product_id': p2.id,
                                    'product_name': p2.name,
                                    'octane': float(p2.octane_number),
                                    'percentage': p2_perc,
                                    'weight_kg': float(float(total_volume) * float(p2_perc) / 100) if total_volume else None,
                                    'price_per_kg': float(p2.price_per_liter)
                                }
                            ],
                            'final_octane': round(final_octane, 2),
                            'final_price_per_kg': round(final_price, 2),
                            'total_price': round(final_price * float(total_volume), 2) if total_volume else None,
                            'gost_compliant': gost_compliant,
                            'gost_warnings': gost_warnings
                        }
                        variants.append(variant)
                
                # Faqat past oktanli productlar bilan variant yaratmaymiz
                # Chunki 3+ product kombinatsiyasi bilan yaxshiroq variant topilishi mumkin
                # (bu variant optimizatsiya funksiyasida topiladi)
    
    # 6. Oktan aniqligini tekshirish va filterlash
    filtered_variants = []
    
    # Oktan diapazonini aniqlash
    product_octanes = [float(p.octane_number) for p in products_list] if products_list else []
    min_oct = min(product_octanes) if product_octanes else 0
    max_oct = max(product_octanes) if product_octanes else 0
    
    # Agar variantlar bo'lsa, barchasini olamiz
    if variants:
        # Oktan farqi bo'yicha tartiblash
        variants.sort(key=lambda x: abs(x['final_octane'] - target_octane))
        
        # Agar maqsad diapazondan tashqarida bo'lsa, eng yaqin variantlarni ko'rsatamiz
        if target_octane > max_oct or target_octane < min_oct:
            max_diff = 100.0  # Juda katta cheklov - barcha variantlarni qabul qilamiz
        else:
            max_diff = 10.0  # Katta cheklov
        
        # Eng yaqin variantlarni filterlash
        for v in variants:
            octane_diff = abs(v['final_octane'] - target_octane)
            if octane_diff <= max_diff:
                filtered_variants.append(v)
                if len(filtered_variants) >= 20:  # Ko'proq variantlar
                    break
    
    # Agar hali ham variantlar topilmasa va maqsad diapazondan tashqarida bo'lsa,
    # maksimal oktan kombinatsiyasini yaratamiz
    if not filtered_variants and products_list and len(products_list) >= 2:
        if target_octane > max_oct:
            # Maksimal oktan kombinatsiyasini yaratish
            from .optimization import solve_maximum_octane
            octanes = [float(p.octane_number) for p in products_list]
            prices = [float(p.price_per_liter) for p in products_list]
            gost_limits = [float(p.gost_percentage or 100) for p in products_list]
            
            max_result = solve_maximum_octane(products_list, octanes, prices, gost_limits)
            if max_result:
                # Format qilish
                variant_products = []
                for i, product in enumerate(max_result['products']):
                    if i < len(max_result['percentages']):
                        pct = max_result['percentages'][i]
                        if pct > 0.01:
                            variant_products.append({
                                'product_id': product.id,
                                'product_name': product.name,
                                'octane': float(product.octane_number),
                                'percentage': round(pct, 2),
                                'weight_kg': float(float(total_volume) * pct / 100) if total_volume else None,
                                'price_per_kg': float(product.price_per_liter)
                            })
                
                if len(variant_products) >= 2:
                    filtered_variants.append({
                        'variant_number': 0,
                        'products': variant_products,
                        'final_octane': max_result['final_octane'],
                        'final_price_per_kg': max_result['final_price'],
                        'total_price': round(max_result['final_price'] * float(total_volume), 2) if total_volume else None,
                        'gost_compliant': max_result.get('gost_compliant', False),
                        'gost_warnings': []
                    })
    
    # 7. Variantlarni tartiblash: GOST talablariga mos keladiganlar birinchi, keyin narx bo'yicha
    if filtered_variants:
        filtered_variants.sort(key=lambda x: (
            not x.get('gost_compliant', False),  # GOSTga mos keladiganlar birinchi (False < True)
            x['final_price_per_kg']  # Keyin narx bo'yicha
        ))
    
    # 7a. Agar kam variantlar bo'lsa (5 dan kam), qo'shimcha variantlar yaratamiz
    if len(filtered_variants) < 5 and products_list and len(products_list) >= 2:
        try:
            from .optimization import optimize_multi_product_blend
            # Turli narx strategiyalari bilan ko'proq variantlar topamiz
            for strategy_multiplier in [0.5, 0.7, 1.0, 1.3, 1.5, 2.0]:
                additional_variants = optimize_multi_product_blend(
                    products_list, 
                    target_octane, 
                    max_products=min(4, len(products_list)),
                    use_ai=True,
                    num_variants=2  # Har bir strategiya uchun 2 ta variant
                )
                
                for ai_var in additional_variants:
                    if not ai_var or not ai_var.get('final_octane'):
                        continue
                    
                    final_octane = ai_var['final_octane']
                    octane_diff = abs(final_octane - target_octane)
                    
                    if octane_diff > 50.0:
                        continue
                    
                    # Variant yaratish
                    variant_products = []
                    for i, product in enumerate(ai_var['products']):
                        if i < len(ai_var['percentages']):
                            pct = ai_var['percentages'][i]
                            if pct > 0.01:
                                variant_products.append({
                                    'product_id': product.id,
                                    'product_name': product.name,
                                    'octane': float(product.octane_number),
                                    'percentage': round(pct, 2),
                                    'weight_kg': float(float(total_volume) * pct / 100) if total_volume else None,
                                    'price_per_kg': float(product.price_per_liter)
                                })
                    
                    if len(variant_products) >= 2:
                        products_perc = [(ai_var['products'][i], ai_var['percentages'][i]) for i in range(len(ai_var['products']))]
                        gost_compliant, gost_warnings = check_gost_compliance(products_perc)
                        
                        new_variant = {
                            'variant_number': 0,
                            'products': variant_products,
                            'final_octane': ai_var['final_octane'],
                            'final_price_per_kg': ai_var['final_price'],
                            'total_price': round(ai_var['final_price'] * float(total_volume), 2) if total_volume else None,
                            'gost_compliant': ai_var.get('gost_compliant', False),
                            'gost_warnings': gost_warnings
                        }
                        
                        # Duplikatlarni tekshirish
                        price_key = round(ai_var['final_price'], 2)
                        existing_prices = {round(v['final_price_per_kg'], 2) for v in filtered_variants}
                        
                        if price_key not in existing_prices:
                            filtered_variants.append(new_variant)
                            if len(filtered_variants) >= 20:  # Yetarli variantlar
                                break
                
                if len(filtered_variants) >= 20:
                    break
            
            # Qayta tartiblash: GOST talablariga mos keladiganlar birinchi, keyin narx bo'yicha
            filtered_variants.sort(key=lambda x: (
                not x.get('gost_compliant', False),  # GOSTga mos keladiganlar birinchi
                x['final_price_per_kg']  # Keyin narx bo'yicha
            ))
        except Exception as e:
            logger.error(f"Qo'shimcha variantlar yaratishda xatolik: {str(e)}", exc_info=True)
    
    # 7b. Agar hali ham variantlar topilmasa, AI algoritmini ishlatamiz
    if not filtered_variants and products_list and len(products_list) >= 2:
        try:
            from .optimization import optimize_multi_product_blend
            ai_variants = optimize_multi_product_blend(
                products_list, 
                target_octane, 
                max_products=min(4, len(products_list)),
                use_ai=True,
                num_variants=10  # Ko'proq variantlar topish uchun
            )
            
            # AI variantlarini format qilish
            for ai_var in ai_variants:
                variant_products = []
                for i, product in enumerate(ai_var['products']):
                    if i < len(ai_var['percentages']):
                        pct = ai_var['percentages'][i]
                        if pct > 0.01:
                            variant_products.append({
                                'product_id': product.id,
                                'product_name': product.name,
                                'octane': float(product.octane_number),
                                'percentage': round(pct, 2),
                                'weight_kg': float(float(total_volume) * pct / 100) if total_volume else None,
                                'price_per_kg': float(product.price_per_liter)
                            })
                
                if len(variant_products) >= 2:
                    filtered_variants.append({
                        'variant_number': 0,
                        'products': variant_products,
                        'final_octane': ai_var['final_octane'],
                        'final_price_per_kg': ai_var['final_price'],
                        'total_price': round(ai_var['final_price'] * float(total_volume), 2) if total_volume else None,
                        'gost_compliant': ai_var.get('gost_compliant', False),
                        'gost_warnings': []
                    })
        except Exception as e:
            logger.error(f"AI optimizatsiya xatoligi: {str(e)}", exc_info=True)
    
    # 8. Kategoriyalarga bo'lish: eng arzon, arzon, o'rtacha, qimmat, juda qimmat
    total_count = len(filtered_variants)
    if total_count == 0:
        return []
    
    # Har bir kategoriyadan eng yaxshi variantni olish
    result_variants = []
    
    # Kategoriya indekslari
    if total_count >= 5:
        indices = [0, total_count // 4, total_count // 2, (total_count * 3) // 4, total_count - 1]
    elif total_count == 4:
        indices = [0, 1, 2, 3, 3]
    elif total_count == 3:
        indices = [0, 1, 2, 2, 2]
    elif total_count == 2:
        indices = [0, 1, 1, 1, 1]
    else:
        indices = [0, 0, 0, 0, 0]
    
    categories = [
        ('eng_arzon', 'Eng arzon'),
        ('arzon', 'Arzon'),
        ('ortacha', 'O\'rtacha'),
        ('qimmat', 'Qimmat'),
        ('juda_qimmat', 'Juda qimmat')
    ]
    
    variant_index = 1
    seen_prices = set()
    
    # Agar kam variantlar bo'lsa (1-2 ta), AI algoritmini bir necha marta chaqirib
    # turli variantlar olamiz
    if total_count < 5 and total_count > 0 and products_list and len(products_list) >= 2:
        # AI algoritmini bir necha marta chaqirib, turli variantlar olamiz
        try:
            from .optimization import optimize_multi_product_blend
            
            # 5 ta variantga yetguncha AI algoritmini bir necha marta chaqiramiz
            # Turli parametrlar bilan turli variantlar yaratamiz
            max_attempts = 20  # Ko'proq urinishlar
            price_strategies = [0.1, 0.3, 0.5, 0.7, 1.0, 1.3, 1.5, 2.0, 2.5, 3.0]  # Turli narx strategiyalari
            
            for strategy_idx, price_weight in enumerate(price_strategies):
                if len(filtered_variants) >= 10:  # Yetarli variantlar
                    break
                    
                # Har bir strategiya uchun bir nechta variant yaratish
                for attempt in range(3):
                    if len(filtered_variants) >= 10:
                        break
                        
                    # optimize_multi_product_blend funksiyasini chaqiramiz
                    # Bu funksiya ichki strategiyalar bilan turli variantlar yaratadi
                    additional_variants = optimize_multi_product_blend(
                        products_list, 
                        target_octane, 
                        max_products=min(4, len(products_list)),
                        use_ai=True,
                        num_variants=1  # Har bir chaqiruv uchun 1 ta variant
                    )
                    
                    for ai_var in additional_variants:
                        if not ai_var or not ai_var.get('final_octane'):
                            continue
                        
                        final_octane = ai_var['final_octane']
                        octane_diff = abs(final_octane - target_octane)
                        
                        # Oktan farqi juda katta bo'lmasligi kerak
                        if octane_diff > 50.0:
                            continue
                        
                        # Variant yaratish
                        variant_products = []
                        for i, product in enumerate(ai_var['products']):
                            if i < len(ai_var['percentages']):
                                pct = ai_var['percentages'][i]
                                if pct > 0.01:
                                    variant_products.append({
                                        'product_id': product.id,
                                        'product_name': product.name,
                                        'octane': float(product.octane_number),
                                        'percentage': round(pct, 2),
                                        'weight_kg': float(float(total_volume) * pct / 100) if total_volume else None,
                                        'price_per_kg': float(product.price_per_liter)
                                    })
                        
                        if len(variant_products) >= 2:
                            products_perc = [(ai_var['products'][i], ai_var['percentages'][i]) for i in range(len(ai_var['products']))]
                            gost_compliant, gost_warnings = check_gost_compliance(products_perc)
                            
                            new_variant = {
                                'variant_number': 0,
                                'products': variant_products,
                                'final_octane': ai_var['final_octane'],
                                'final_price_per_kg': ai_var['final_price'],
                                'total_price': round(ai_var['final_price'] * float(total_volume), 2) if total_volume else None,
                                'gost_compliant': ai_var.get('gost_compliant', False),
                                'gost_warnings': gost_warnings
                            }
                            
                            # Duplikatlarni tekshirish - narx va kompozitsiya bo'yicha
                            price_key = round(ai_var['final_price'], 2)
                            existing_prices = {round(v['final_price_per_kg'], 2) for v in filtered_variants}
                            
                            # Kompozitsiyani ham tekshirish
                            is_duplicate = False
                            for existing_variant in filtered_variants:
                                existing_composition = sorted([(p['product_id'], round(p['percentage'], 1)) for p in existing_variant['products']])
                                new_composition = sorted([(p['product_id'], round(p['percentage'], 1)) for p in variant_products])
                                if existing_composition == new_composition:
                                    is_duplicate = True
                                    break
                            
                            if not is_duplicate and price_key not in existing_prices:
                                filtered_variants.append(new_variant)
                                if len(filtered_variants) >= 10:
                                    break
            
            # Qayta tartiblash
            if filtered_variants:
                # GOST talablariga mos keladiganlar birinchi, keyin narx bo'yicha
                filtered_variants.sort(key=lambda x: (
                    not x.get('gost_compliant', False),  # GOSTga mos keladiganlar birinchi
                    x['final_price_per_kg']  # Keyin narx bo'yicha
                ))
                total_count = len(filtered_variants)
        except Exception as e:
            logger.error(f"Qo'shimcha variantlar yaratishda xatolik: {str(e)}", exc_info=True)
        
        # Barcha variantlarni ko'rsatamiz (yoki yangi yaratilganlar bilan)
        for idx, variant in enumerate(filtered_variants[:5]):
            variant = variant.copy()
            category_name, category_label = categories[min(idx, len(categories) - 1)]
            variant['variant_number'] = variant_index
            variant['category'] = category_name
            variant['category_label'] = category_label
            result_variants.append(variant)
            variant_index += 1
    else:
        # 5 ta kategoriyadan variantlarni olamiz
        for idx, (category_name, category_label) in zip(indices, categories):
            if 0 <= idx < total_count:
                variant = filtered_variants[idx].copy()
                price_key = round(variant['final_price_per_kg'], 2)
                
                # Agar bir xil narxli variant bo'lmasa qo'shamiz
                if price_key not in seen_prices:
                    variant['variant_number'] = variant_index
                    variant['category'] = category_name
                    variant['category_label'] = category_label
                    result_variants.append(variant)
                    seen_prices.add(price_key)
                    variant_index += 1
    
    return result_variants[:5]  # Maksimal 5 ta variant


def product_selection(request):
    """Страница выбора продуктов и ввода параметров"""
    # Bazadagi barcha aralashma uchun productlarni olamiz
    products = Product.objects.filter(
        is_for_blending=True
    ).order_by('name')
    
    return render(request, 'calibration/product_selection.html', {
        'products': products
    })


def gasoline_blend_calculator(request):
    """Страница выбора октана и расчета"""
    return render(request, 'calibration/gasoline_blend_calculator.html')


@csrf_exempt
@require_http_methods(["POST"])
def calculate_gasoline_blend(request):
    """AJAX endpoint: benzin aralashma hisob-kitoblari"""
    try:
        data = json.loads(request.body)
        
        target_octane_str = str(data.get('target_octane', '')).strip()
        total_weight_str = str(data.get('total_weight', '') or data.get('total_volume', '')).strip()  # Frontend dan "total_weight" yoki "total_volume" keladi
        variants_count = int(data.get('variants_count', 5))
        products_data = data.get('products', {})  # {product_id: {octane, price, gost_percentage}}
        
        # Validatsiya
        if not target_octane_str:
            return JsonResponse({
                'success': False,
                'error': 'Введите октановое число'
            })
        
        try:
            target_octane = int(target_octane_str)
            if target_octane <= 0:
                raise ValueError()
        except (ValueError, TypeError):
            return JsonResponse({
                'success': False,
                'error': 'Неверное октановое число'
            })
        
        # Products validatsiya
        if not products_data:
            return JsonResponse({
                'success': False,
                'error': 'Введите октановое число и цену хотя бы для 2 продуктов'
            })
        
        # Valid productslarni filtrlash
        valid_products = {}
        for product_id, p_data in products_data.items():
            try:
                octane = float(p_data.get('octane', 0))
                price = float(p_data.get('price', 0))
                if octane > 0 and price > 0:
                    valid_products[product_id] = {
                        'octane': octane,
                        'price': price,
                        'gost_percentage': float(p_data.get('gost_percentage', 100)) if p_data.get('gost_percentage') else 100
                    }
            except (ValueError, TypeError):
                continue
        
        if len(valid_products) < 2:
            return JsonResponse({
                'success': False,
                'error': 'Введите октановое число и цену хотя бы для 2 продуктов'
            })
        
        total_weight = None
        if total_weight_str:
            try:
                total_weight = Decimal(total_weight_str)
                if total_weight <= 0:
                    total_weight = None
            except (ValueError, TypeError):
                total_weight = None
        
        # Variantlarni hisoblash
        # total_weight aslida kg, lekin find_blend_variants funksiyasi total_volume parametrini kutadi
        try:
            variants = find_blend_variants(target_octane, valid_products, max_variants=variants_count, total_volume=total_weight)
            
            logger.info(f"Topilgan variantlar soni: {len(variants) if variants else 0}")
            logger.info(f"Maqsad oktan: {target_octane}, Productlar: {list(valid_products.keys())}")
            
            if not variants:
                # Debug ma'lumotlari
                product_octanes = [p.get('octane', 0) for p in valid_products.values()]
                min_oct = min(product_octanes) if product_octanes else 0
                max_oct = max(product_octanes) if product_octanes else 0
                
                # Agar maqsad diapazondan tashqarida bo'lsa, ham eng yaqin variantlarni ko'rsatamiz
                # Bu holda, maksimal oktan kombinatsiyasini qaytaramiz
                logger.warning(f"Variantlar topilmadi. Maqsad: {target_octane}, Diapazon: {min_oct}-{max_oct}")
                
                # Yana bir bor optimizatsiyani sinab ko'ramiz - AI algoritmi bilan
                try:
                    # AI algoritmi ishlatish
                    from .optimization import optimize_multi_product_blend
                    from .models import Product
                    
                    # Product objectlarini olamiz
                    product_objects = []
                    for prod_id, prod_data in valid_products.items():
                        try:
                            product = Product.objects.get(id=prod_id)
                            # Oktan va narxni frontend'dan o'rnatamiz
                            if prod_data.get('octane'):
                                product.octane_number = prod_data['octane']
                            if prod_data.get('price'):
                                product.price_per_liter = prod_data['price']
                            product_objects.append(product)
                        except Product.DoesNotExist:
                            continue
                    
                    if len(product_objects) >= 3:
                        ai_variants = optimize_multi_product_blend(
                            product_objects, 
                            target_octane, 
                            max_products=min(4, len(product_objects)),
                            use_ai=True,
                            num_variants=3
                        )
                        
                        if ai_variants:
                            # AI variantlarini format qilamiz
                            formatted_variants = []
                            for ai_var in ai_variants[:5]:  # Maksimal 5 ta
                                variant_products = []
                                for i, product in enumerate(ai_var['products']):
                                    if i < len(ai_var['percentages']):
                                        pct = ai_var['percentages'][i]
                                        if pct > 0.01:
                                            variant_products.append({
                                                'product_id': product.id,
                                                'product_name': product.name,
                                                'octane': float(product.octane_number),
                                                'percentage': round(pct, 2),
                                                'weight_kg': float(total_weight * pct / 100) if total_weight else None,
                                                'price_per_kg': float(product.price_per_liter)
                                            })
                                
                                if len(variant_products) >= 2:
                                    formatted_variants.append({
                                        'variant_number': 0,
                                        'products': variant_products,
                                        'final_octane': ai_var['final_octane'],
                                        'final_price_per_kg': ai_var['final_price'],
                                        'total_price': round(ai_var['final_price'] * float(total_weight), 2) if total_weight else None,
                                        'gost_compliant': ai_var.get('gost_compliant', False),
                                        'gost_warnings': []
                                    })
                            
                            if formatted_variants:
                                # Variantlarni format qilish va qaytarish
                                categorized_variants = categorize_variants_by_price(formatted_variants)
                                
                                return JsonResponse({
                                    'success': True,
                                    'variants': categorized_variants,
                                    'message': f'AI algoritmi {len(categorized_variants)} ta variant topdi (eng yaqin)'
                                })
                except Exception as e:
                    logger.error(f"AI optimizatsiya xatoligi: {str(e)}", exc_info=True)
                
                # Agar hali ham variantlar topilmasa
                return JsonResponse({
                    'success': False,
                    'error': f'AI-{target_octane} uchun ideal variantlar topilmadi. Mavjud productlar oktan diapazoni: {min_oct}-{max_oct}. Maqsad oktan: {target_octane}. Iltimos, maqsad oktan sonini {min_oct}-{max_oct} orasida tanlang yoki yuqori oktanli productlar qo\'shing.'
                })
        except Exception as e:
            logger.error(f"Variantlarni topishda xatolik: {str(e)}", exc_info=True)
            return JsonResponse({
                'success': False,
                'error': f'Variantlarni topishda xatolik: {str(e)}'
            })
        
        # Variantlarni raqamlash
        for i, variant in enumerate(variants, 1):
            variant['variant_number'] = i
        
        # Eng yaxshi variant indeksini topish (eng arzon)
        best_index = 0
        
        # Bazaga saqlash
        calculation = GasolineBlendCalculation.objects.create(
            target_octane=target_octane,
            total_volume_liters=total_weight,  # Bu aslida kg, lekin modelda hali "liters" nomi
            variants_count=len(variants),
            blend_variants=variants,
            best_variant_index=best_index,
            calculation_method='linear'
        )
        
        return JsonResponse({
            'success': True,
            'calculation_id': calculation.id,
            'target_octane': target_octane,
            'variants': variants,
            'variants_count': len(variants),
            'best_variant_index': best_index
        })
        
    except Exception as e:
        import traceback
        logger.error(f"Gasoline blend calculation error: {str(e)}")
        logger.error(traceback.format_exc())
        return JsonResponse({
            'success': False,
            'error': f'Ошибка: {str(e)}'
        })


@csrf_exempt
@require_http_methods(["POST"])
def save_product_configuration(request):
    """AJAX endpoint: product konfiguratsiyasini saqlash"""
    try:
        data = json.loads(request.body)
        
        name = data.get('name', 'Konfiguratsiya').strip()
        products_config = data.get('products', {})
        description = data.get('description', '').strip()
        
        if not name:
            name = f"Konfiguratsiya {SavedProductConfiguration.objects.count() + 1}"
        
        if not products_config:
            return JsonResponse({
                'success': False,
                'error': 'Конфигурация продуктов пуста'
            })
        
        # Saqlash
        config = SavedProductConfiguration.objects.create(
            name=name,
            products_config=products_config,
            description=description if description else None,
            is_active=True
        )
        
        return JsonResponse({
            'success': True,
            'config_id': config.id,
            'message': 'Конфигурация успешно сохранена!'
        })
        
    except Exception as e:
        logger.error(f"Save configuration error: {str(e)}")
        return JsonResponse({
            'success': False,
            'error': f'Ошибка: {str(e)}'
        })


def saved_configurations_list(request):
    """Список сохраненных конфигураций"""
    configs = SavedProductConfiguration.objects.filter(is_active=True).order_by('-created_at')
    
    # Product ma'lumotlarini yuklash
    configs_with_products = []
    for config in configs:
        products_info = []
        for product_id, p_data in config.products_config.items():
            try:
                product = Product.objects.get(id=int(product_id))
                products_info.append({
                    'product': product,
                    'octane': p_data.get('octane'),
                    'price': p_data.get('price'),
                    'gost_percentage': p_data.get('gost_percentage')
                })
            except (Product.DoesNotExist, ValueError, TypeError):
                continue
        
        configs_with_products.append({
            'config': config,
            'products': products_info
        })
    
    return render(request, 'calibration/saved_configurations.html', {
        'configs_with_products': configs_with_products
    })


def load_configuration(request, config_id):
    """Загрузка сохраненной конфигурации (возвращает JSON)"""
    try:
        config = get_object_or_404(SavedProductConfiguration, id=config_id, is_active=True)
        
        return JsonResponse({
            'success': True,
            'name': config.name,
            'products': config.products_config,
            'description': config.description
        })
    except Exception as e:
        logger.error(f"Load configuration error: {str(e)}")
        return JsonResponse({
            'success': False,
            'error': f'Ошибка: {str(e)}'
        })


@require_http_methods(["POST"])
@csrf_exempt
def delete_configuration(request, config_id):
    """Удаление сохраненной конфигурации"""
    try:
        config = get_object_or_404(SavedProductConfiguration, id=config_id, is_active=True)
        config.is_active = False  # Soft delete
        config.save()
        
        return JsonResponse({
            'success': True,
            'message': 'Конфигурация успешно удалена'
        })
    except Exception as e:
        logger.error(f"Delete configuration error: {str(e)}")
        return JsonResponse({
            'success': False,
            'error': f'Ошибка: {str(e)}'
        })


@csrf_exempt
@require_http_methods(["POST"])
def export_blend_variants_excel(request):
    """Excel ga variantlarni eksport qilish"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from django.http import HttpResponse
        import json
        from datetime import datetime
        
        data = json.loads(request.body)
        variants = data.get('variants', [])
        target_octane = data.get('target_octane', 'N/A')
        total_weight = data.get('total_weight', None)
        
        if not variants:
            return JsonResponse({
                'success': False,
                'error': 'Eksport qilish uchun variantlar mavjud emas'
            })
        
        # Workbook yaratish
        wb = Workbook()
        ws = wb.active
        ws.title = "Benzin Aralashma Variantlari"
        
        # Uslublar
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        category_colors = {
            'eng_arzon': PatternFill(start_color="28A745", end_color="28A745", fill_type="solid"),  # Yashil
            'arzon': PatternFill(start_color="17A2B8", end_color="17A2B8", fill_type="solid"),  # Ko'k
            'ortacha': PatternFill(start_color="FFC107", end_color="FFC107", fill_type="solid"),  # Sariq
            'qimmat': PatternFill(start_color="FF8C00", end_color="FF8C00", fill_type="solid"),  # Qizg'ish
            'juda_qimmat': PatternFill(start_color="DC3545", end_color="DC3545", fill_type="solid"),  # Qizil
        }
        
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Sarlavha
        ws.merge_cells('A1:F1')
        title_cell = ws['A1']
        title_cell.value = f"Benzin Aralashma Kalkulyatori - AI-{target_octane}"
        title_cell.font = Font(bold=True, size=14)
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        title_cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        
        # Sana va vaqt
        ws.merge_cells('A2:F2')
        date_cell = ws['A2']
        date_cell.value = f"Tayyorlangan: {datetime.now().strftime('%d.%m.%Y %H:%M')}"
        date_cell.font = Font(size=10, italic=True)
        date_cell.alignment = Alignment(horizontal='center')
        
        # Umumiy ma'lumotlar
        if total_weight:
            ws.merge_cells('A3:F3')
            info_cell = ws['A3']
            info_cell.value = f"Umumiy og'irlik: {total_weight} kg"
            info_cell.font = Font(size=10)
            info_cell.alignment = Alignment(horizontal='center')
            start_row = 5
        else:
            start_row = 4
        
        # Har bir variant uchun
        current_row = start_row
        
        for variant in variants:
            category = variant.get('category', 'ortacha')
            category_label = variant.get('category_label', 'O\'rtacha')
            variant_number = variant.get('variant_number', 0)
            final_octane = variant.get('final_octane', 0)
            final_price_per_kg = variant.get('final_price_per_kg', variant.get('final_price_per_liter', 0))
            total_price = variant.get('total_price', None)
            gost_compliant = variant.get('gost_compliant', False)
            products = variant.get('products', [])
            
            # Variant sarlavhasi
            variant_start_row = current_row
            
            # Kategoriya va variant raqami
            ws.merge_cells(f'A{current_row}:F{current_row}')
            header_cell = ws[f'A{current_row}']
            header_cell.value = f"Variant {variant_number} - {category_label}"
            header_cell.font = header_font
            header_cell.fill = category_colors.get(category, category_colors['ortacha'])
            header_cell.alignment = Alignment(horizontal='center', vertical='center')
            header_cell.border = border
            current_row += 1
            
            # Asosiy ma'lumotlar
            ws[f'A{current_row}'] = "Oktan soni:"
            ws[f'A{current_row}'].font = Font(bold=True)
            ws[f'B{current_row}'] = f"{final_octane} (maqsad: {target_octane})"
            
            ws[f'C{current_row}'] = "Narx (so'm/kg):"
            ws[f'C{current_row}'].font = Font(bold=True)
            ws[f'D{current_row}'] = f"{final_price_per_kg:,.2f}"
            
            if total_price:
                ws[f'E{current_row}'] = "Jami narx:"
                ws[f'E{current_row}'].font = Font(bold=True)
                ws[f'F{current_row}'] = f"{total_price:,.2f} so'm"
            
            for col in ['A', 'B', 'C', 'D', 'E', 'F']:
                ws[f'{col}{current_row}'].border = border
            current_row += 1
            
            # GOST ma'lumoti
            ws[f'A{current_row}'] = "GOST:"
            ws[f'A{current_row}'].font = Font(bold=True)
            ws[f'B{current_row}'] = "✅ Mos keladi" if gost_compliant else "⚠️ Mos kelmaydi"
            ws[f'B{current_row}'].font = Font(color="28A745" if gost_compliant else "DC3545")
            for col in ['A', 'B', 'C', 'D', 'E', 'F']:
                ws[f'{col}{current_row}'].border = border
            current_row += 1
            
            # Kompozitsiya jadval sarlavhasi
            ws[f'A{current_row}'] = "Product nomi"
            ws[f'B{current_row}'] = "Oktan"
            ws[f'C{current_row}'] = "Foiz (%)"
            ws[f'D{current_row}'] = "Narx (so'm/kg)"
            if total_weight:
                ws[f'E{current_row}'] = "Og'irlik (kg)"
            ws[f'F{current_row}'] = "Jami narx"
            
            for col in ['A', 'B', 'C', 'D', 'E', 'F']:
                cell = ws[f'{col}{current_row}']
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border
            current_row += 1
            
            # Kompozitsiya
            for product in products:
                ws[f'A{current_row}'] = product.get('product_name', '')
                ws[f'B{current_row}'] = product.get('octane', 0)
                ws[f'C{current_row}'] = product.get('percentage', 0)
                ws[f'D{current_row}'] = product.get('price_per_kg', product.get('price_per_liter', 0))
                
                if total_weight:
                    weight = product.get('weight_kg', product.get('volume_liters', 0))
                    ws[f'E{current_row}'] = weight if weight else ''
                
                # Jami narx (foiz * jami og'irlik * narx)
                if total_weight and product.get('percentage'):
                    product_total = (float(total_weight) * product.get('percentage', 0) / 100) * product.get('price_per_kg', product.get('price_per_liter', 0))
                    ws[f'F{current_row}'] = f"{product_total:,.2f}"
                
                for col in ['A', 'B', 'C', 'D', 'E', 'F']:
                    cell = ws[f'{col}{current_row}']
                    cell.border = border
                    if col in ['B', 'C', 'D', 'E', 'F']:
                        cell.alignment = Alignment(horizontal='right', vertical='center')
                    else:
                        cell.alignment = Alignment(horizontal='left', vertical='center')
                
                current_row += 1
            
            # Bo'sh qator
            current_row += 1
        
        # Ustunlarni kenglashtirish
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 18
        ws.column_dimensions['E'].width = 18
        ws.column_dimensions['F'].width = 18
        
        # HTTP response yaratish
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f"benzin_aralashma_AI{target_octane}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        wb.save(response)
        return response
        
    except json.JSONDecodeError as e:
        logger.error(f"JSON decode xatoligi: {str(e)}")
        return JsonResponse({
            'success': False,
            'error': f'Noto\'g\'ri JSON format: {str(e)}'
        }, status=400)
    except json.JSONDecodeError as e:
        logger.error(f"JSON decode xatoligi: {str(e)}")
        return JsonResponse({
            'success': False,
            'error': f'Noto\'g\'ri JSON format: {str(e)}'
        }, status=400)
    except Exception as e:
        logger.error(f"Excel export xatoligi: {str(e)}", exc_info=True)
        import traceback
        error_trace = traceback.format_exc()
        logger.error(f"Full traceback:\n{error_trace}")
        
        # Xatolikni JSON formatda qaytaramiz (HTML emas)
        return JsonResponse({
            'success': False,
            'error': f'Excel eksport xatoligi: {str(e)}',
            'traceback': error_trace if logger.level <= logging.DEBUG else None
        }, status=500)


def compare_variants(request):
    """Variantlarni solishtirish sahifasi"""
    return render(request, 'calibration/compare_variants.html')


def view_gasoline_blend_history(request, calculation_id):
    """Tarixdagi gasoline blend calculationni ko'rish"""
    try:
        calculation = get_object_or_404(GasolineBlendCalculation, id=calculation_id)
        
        # Calculation ma'lumotlarini formatlash
        variants = calculation.blend_variants if calculation.blend_variants else []
        target_octane = calculation.target_octane
        total_weight = float(calculation.total_volume_liters) if calculation.total_volume_liters else None
        
        # JSONField dan olingan ma'lumotlar allaqachon dict/list bo'ladi
        # JSON formatda template'ga uzatish uchun
        variants_json = json.dumps(variants, ensure_ascii=False, default=str)
        
        # Template'ga uzatish uchun ma'lumotlar
        context = {
            'calculation': calculation,
            'variants_json': variants_json,  # JSON string sifatida JavaScript uchun
            'variants': variants,  # Python list/dict formatida ham (agar kerak bo'lsa)
            'target_octane': target_octane,
            'total_weight': total_weight,
            'variants_count': len(variants),
            'best_variant_index': calculation.best_variant_index,
            'calculation_date': calculation.timestamp.strftime('%d.%m.%Y %H:%M'),
        }
        
        return render(request, 'calibration/view_gasoline_blend_history.html', context)
        
    except GasolineBlendCalculation.DoesNotExist:
        messages.error(request, 'Hisoblash topilmadi')
        return redirect('calibration:history')
    except Exception as e:
        logger.error(f"Gasoline blend history view error: {str(e)}", exc_info=True)
        messages.error(request, f'Xatolik: {str(e)}')
        return redirect('calibration:history')


def calculator_selector(request):
    """Страница выбора калькулятора"""
    return render(request, 'calibration/calculator_selector.html')
